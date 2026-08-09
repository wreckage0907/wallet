# Copyright (c) 2026, girish.raghav2004@gmail.com and contributors
# For license information, please see license.txt

"""Turning a statement file into a plain grid of cells.

Everything downstream (header detection, column mapping, row parsing) works on
`list[list[cell]]`, so this is the only module that knows about file formats.
"""

import csv
import io

import frappe
from frappe import _

from wallet.statement.decrypt import decrypt

#: Statements longer than this are almost certainly not statements.
MAX_SCAN_ROWS = 100_000


def read_grid(
	content: bytes, file_name: str | None = None, password: str | None = None, sheet_name: str | None = None
) -> tuple[list[list], str]:
	"""Return (grid, sheet_name) for a statement file.

	Cells keep their native types where the format has them: openpyxl hands back real
	`datetime` and `float` objects, which is strictly better than re-parsing strings.
	"""
	content = decrypt(content, password)
	extension = (file_name or "").rsplit(".", 1)[-1].lower()

	if extension == "csv" or (not extension and content[:2] != b"PK"):
		return _read_csv(content), "csv"

	if extension == "xls" and content[:2] != b"PK":
		return _read_xls(content, sheet_name)

	return _read_xlsx(content, sheet_name)


def _read_xlsx(content: bytes, sheet_name: str | None = None) -> tuple[list[list], str]:
	import openpyxl

	try:
		workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	except Exception as e:
		frappe.throw(_("Could not read this file as an Excel workbook: {0}").format(e))

	try:
		sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
		grid = [list(row) for row in sheet.iter_rows(values_only=True, max_row=MAX_SCAN_ROWS)]
		return grid, sheet.title
	finally:
		workbook.close()


def _read_xls(content: bytes, sheet_name: str | None = None) -> tuple[list[list], str]:
	import xlrd

	book = xlrd.open_workbook(file_contents=content)
	sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)

	# row_values() hands back date cells as raw Excel serial floats, so a legacy .xls
	# whose dates are real date-typed cells would stage zero transactions - every row
	# would look undated. Convert them using the workbook's epoch.
	grid = []
	for r in range(sheet.nrows):
		row = []
		for c in range(sheet.ncols):
			cell = sheet.cell(r, c)
			if cell.ctype == xlrd.XL_CELL_DATE:
				try:
					row.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
					continue
				except (ValueError, OverflowError):
					pass
			row.append(cell.value)
		grid.append(row)

	return grid, sheet.name


def _read_csv(content: bytes) -> list[list]:
	text = content.decode("utf-8-sig", errors="replace")
	# Banks emit comma, semicolon and tab separated files interchangeably.
	try:
		dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
	except csv.Error:
		dialect = csv.excel

	return [list(row) for row in csv.reader(io.StringIO(text), dialect)]
